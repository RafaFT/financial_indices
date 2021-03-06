from abc import (ABCMeta,
                 abstractmethod)
import datetime
import decimal
import logging
import os
from types import MappingProxyType
from typing import (Dict,
                    Collection,
                    Iterable,
                    Optional,
                    Set,
                    Tuple,
                    Union,
                    )

import openpyxl as xlsx

from bcb_api import (DAY_RECORD,
                     IndicatorRecord,
                     RECORDS)
import utils


logger = logging.getLogger('__main__.' + __name__)


class WorksheetWriter(metaclass=ABCMeta):
    """ Class responsible for writing information in a Worksheet object."""

    def __init__(self, worksheet: 'openpyxl.worksheet.worksheet.Worksheet',
                 records: RECORDS):
        """
        :param worksheet:
        :param records:
        """

        self._worksheet = worksheet
        self._indicator_records = records
        self._headers = self._get_headers()

        self._write_records()

    def _get_headers(self) -> Tuple[str]:
        """ Return the name of the headers. Headers may be text or numbers.
        """

        return ('date', 'value',)

    @abstractmethod
    def _format_record(self, record: DAY_RECORD) -> Collection:
        """ Return a collection of values corresponding to a row of data."""
        pass

    def _write_headers(self) -> None:
        """ Write the self._headers values starting at row 1, column 1."""

        for column, header in enumerate(self._headers, 1):
            self._worksheet.cell(1, column).value = header

    def _get_first_row(self, first_date: datetime.date) -> int:
        """ Return the first row to start writing the self._indicators_records,
        based on the first_date value provided.

        :param first_date: The first date from self._indicators_records.
        :return: Integer of the row to start writing.
        """

        column = 1
        for row in range(self._worksheet.max_row, 1, -1):
            date_on_cell = self._worksheet.cell(row, column).value.date()

            if date_on_cell == first_date:
                return row
            elif date_on_cell < first_date:
                return row + 1
        else:
            return 1

    def _erase_extra_records(self, row: int) -> None:
        """ Removes all records from row to the max row, and column 1 to the
        max column.

        :param row: Integer of the first row to erase values.
        :return: None.
        """

        for row in range(row, self._worksheet.max_row + 1):
            for column in range(1, self._worksheet.max_column + 1):
                self._worksheet.cell(row, column).value = None

    def _write_records(self) -> None:
        """ Write all dates and values from self._indicators_records in
        self._worksheet.

        :return: None.
        """

        try:
            first_date = self._indicator_records[0].date
        except KeyError:
            first_row = 1
        else:
            first_row = self._get_first_row(first_date)

        if first_row == 1:
            self._write_headers()
            first_row = 2

        for row, record in enumerate(self._indicator_records, first_row):
            formatted_record = self._format_record(record)
            for column, column_data in enumerate(formatted_record, 1):
                self._worksheet.cell(row, column).value = column_data

        try:
            self._erase_extra_records(row + 1)
        except NameError:
            self._erase_extra_records(first_row)


class SelicWriter(WorksheetWriter):

    def _get_headers(self) -> Tuple[str]:
        """ Generate the headers for a Selic worksheet.

        :return: Tuple of string.
        """

        return super()._get_headers() + ('daily_value',)

    def _format_record(self, record: DAY_RECORD
                       ) -> Tuple[datetime.date, decimal.Decimal, decimal.Decimal]:
        """ Format a record to be appropriate to the worksheet selic.

        :param record: IndicatorRecord.
        :return: Iterable of the values of record.
        """

        daily_value = (1 + round(record.value * 1 / 100, 8))

        return (record.date, record.value, daily_value)


class CdiWriter(WorksheetWriter):

    def _get_headers(self) -> Tuple[str]:
        """ Generate the headers for a CDI worksheet.

        :return: Tuple of string.
        """

        return super()._get_headers() + ('daily_value',)

    def _format_record(self, record: DAY_RECORD
                       ) -> Tuple[datetime.date, decimal.Decimal, decimal.Decimal]:
        """ Format a record to be appropriate to the worksheet cdi.

        :param record: IndicatorRecord.
        :return: Iterable of the values of record.
        """

        daily_value = (1 + round(record.value * 1 / 100, 8))

        return (record.date, record.value, daily_value)


class IpcaWriter(WorksheetWriter):

    def _get_headers(self) -> Tuple[str]:
        """ Return the name of the headers. Headers may be string or numbers.
        """

        return ('ano', 'mes', 'valor')

    def _format_record(self, record: DAY_RECORD
                       ) -> Iterable[Union[int, decimal.Decimal]]:
        """ Format a record to be appropriate to the worksheet ipca.

        :param record: IndicatorRecord.
        :return: Iterable of the values of record.
        """

        return record.date.year, record.date.month, record.value

    def _get_first_row(self, first_date: datetime.date) -> int:
        """ Return the first row to start writing the self._indicators_records,
        based on the first_date value provided.

        :param first_date: The first date from self._indicators_records.
        :return: Integer of the row to start writing.
        """

        for row in range(self._worksheet.max_row, 1, -1):
            year = self._worksheet.cell(row, 1).value
            month = self._worksheet.cell(row, 2).value
            date_on_cell = datetime.date(year, month, 1)

            if date_on_cell == first_date:
                return row
            elif date_on_cell < first_date:
                return row + 1
        else:
            return 1


class TrWriter(WorksheetWriter):

    def _get_headers(self) -> Tuple[str]:
        """ Return the name of the headers. Headers may be string or numbers.
        """

        return ('data inicial', 'data final', 'valor')

    def _format_record(self, record: DAY_RECORD) -> Collection:
        """ Format a record to be appropriate to the worksheet tr.

        :param record: IndicatorRecord.
        :return: Iterable of the values of record.
        """

        return record.date, record.end_date, record.value


class MetadataWriter:
    """ Class to write and update values in the worksheet responsible for storing
    metadata information about each indicator.
    """

    def __init__(self, worksheet: 'openpyxl.worksheet.worksheet.Worksheet'):
        """ Constructor of MetadataWriter."""

        self._worksheet = worksheet

        self._write_headers()
        self.indicators_dates = self._get_indicator_last_date()

    def _write_headers(self) -> None:
        """ Write the the header values starting at row 1, column 1."""

        for column, header in enumerate(('indicator', 'last date'), 1):
            self._worksheet.cell(1, column).value = header

    def _get_indicator_last_date(self) -> Dict[int, Union[datetime.date, None]]:
        """ Return a dictionary with all existing indicators and last dates stored
        in self._worksheet.
        """
        indicador_date = {}
        for row in range(self._worksheet.max_row, 1, -1):
            try:
                cod = int(self._worksheet.cell(row, 1).value)
            except TypeError:
                continue
            try:
                date = self._worksheet.cell(row, 2).value.date()
            except AttributeError:
                date = None
            indicador_date[cod] = date

        return indicador_date

    def write_indicators_last_date(self) -> None:
        """ Writes self.indicators_dates values on self._worksheet."""

        logger.info('Updating Metadata information.')
        row = 2
        for indicator, date in sorted(self.indicators_dates.items()):
            logger.debug(f'New latest date for: {indicator} -> {date}')
            self._worksheet.cell(row, 1).value = indicator
            self._worksheet.cell(row, 2).value = date
            row += 1


class IndicatorsWorkbook:
    """ Class to represent an excel Workbook."""

    _worksheet_protection = MappingProxyType({
        'autoFilter': True,
        'deleteColumns': True,
        'deleteRows': True,
        'formatCells': True,
        'formatColumns': False,
        'formatRows': False,
        'insertColumns': True,
        'insertHyperlinks': True,
        'insertRows': True,
        'objects': False,
        'password': 'KawhiLeonardRocks',
        'pivotTables': True,
        'scenarios': False,
        'selectLockedCells': False,
        'selectUnlockedCells': False,
        'sheet': True,
        'sort': True,
    }
    )

    _worksheet_properties = MappingProxyType(
        {
            -1: {
                'name': 'metadata',
                'color': '000000',
                'writer': MetadataWriter,
                'state': 'veryHidden',
            },
            11: {
                'name': 'selic',
                'color': '0000FF',  # blue
                'writer': SelicWriter,
                'state': 'visible',
            },
            12: {
                'name': 'cdi',
                'color': '00FF00',  # green
                'writer': CdiWriter,
                'state': 'visible',
            },
            433: {
                'name': 'ipca',
                'color': 'FFA500',  # orange
                'writer': IpcaWriter,
                'state': 'visible',
            },
            226: {
                'name': 'tr',
                'color': 'FF0000',  # red
                'writer': TrWriter,
                'state': 'visible',
            },
        },
    )

    def __init__(self, path_to_file: Optional[str] = None,
                 filename: str = 'financial-indicators.xlsx') -> None:
        """ Constructor of a workbook.
        If path_to_file is None, than it is set to the current working directory.
        If filename exists in path_to_file, it is loaded, otherwise a new file
        is created.

        :param path_to_file: String of a valid path, where the filename exists.
        :param filename: Name of the file that either is being load or created.
        """

        if path_to_file is None:
            # path_to_file value depends if the program is being run by
            # a python interpretor or as an executable.
            path_to_file = utils.bundle_dir

        self._workbook_path = os.path.join(path_to_file, filename)

        if filename in os.listdir(path_to_file):
            logger.info(f'Loading workbook: {self._workbook_path}')
            self._workbook = xlsx.load_workbook(self._workbook_path)
        else:
            logger.info(f'Creating new workbook: {self._workbook_path}')
            self._workbook = xlsx.Workbook()
            self._delete_all_sheets()

        worksheet_metadata = self._create_sheet(-1)
        metadata_writer = self.__class__._worksheet_properties[-1]['writer']
        self._metadata_writer = metadata_writer(worksheet_metadata)

    def __len__(self):
        """ Return the number of worksheets inside self._workbook."""

        return len(self._workbook.sheetnames)

    def __repr__(self):
        return f'{self.__class__.__name__}({self._workbook_path})'

    def _delete_all_sheets(self) -> None:
        """ Delete all existing worksheets from self._workbook."""

        for sheet in self._workbook.sheetnames:
            logger.debug(f'Erasing worksheet="{sheet}"')
            del self._workbook[sheet]

    def _create_sheet(self, indicators_code: int
                      ) -> 'openpyxl.worksheet.worksheet.Worksheet':
        """ Create and return a worksheet in self._workbook based on the
        indicators_code value. If that indicators_code worksheet already exists, it
        is simply returned.

        :param indicators_code: Integer representing a financial indicator.
        :return: Worksheet object.
        """

        name = self.__class__._worksheet_properties[indicators_code]['name']
        try:
            ws = self._workbook[name]
            logger.info(f'Loaded worksheet {name}')
            return ws
        except KeyError:
            color = self.__class__._worksheet_properties[indicators_code]['color']
            state = self.__class__._worksheet_properties[indicators_code]['state']

            ws = self._workbook.create_sheet(name)
            ws.title = name
            ws.sheet_properties.tabColor = color
            ws.sheet_state = state

            logger.info(f'Created worksheet {name}')

            return ws

    def get_indicator_last_date(self, indicator_code: int) -> Optional[datetime.date]:
        """ Return the date of indicator_code on the self._metadata_writer.
        If indicator_code value is not present in self._metadata_writer, None
        is returned.

        :param indicator_code: Integer representing a financial indicator.
        :return: Date or None.
        """

        try:
            lastest_date = self._metadata_writer.indicators_dates[indicator_code]
        except KeyError:
            lastest_date = None

        logger.info(f'Latest date of {indicator_code} is {lastest_date}')

        return lastest_date

    @utils.log_func_time(logger, 20)
    def write_records(self, indicator_code: int,
                      records: RECORDS,
                      last_non_extended_date: Optional[datetime.date] = None
                      ) -> None:
        """ Create or load a worksheet from self._workbook, corresponding to
        the indicator_code provided, and pass both the worksheet and records
        to the correct WorksheetWriter (ex: CdiWriter).

        :param indicator_code: Integer representing a financial indicator.
        :param records: Records of a financial indicator.
        :param last_non_extended_date: The last date on records, that is a real
            record, and not an extended one.
        :return: None.
        """

        name = self.__class__._worksheet_properties[indicator_code]['name']
        writer = self.__class__._worksheet_properties[indicator_code]['writer']

        ws = self._create_sheet(indicator_code)

        logger.info(
            f'''Writing {len(records)} record(s) to sheet="{name}"" with last date
                as {last_non_extended_date}
                ''')

        writer(ws, records)
        self._metadata_writer.indicators_dates[indicator_code] = last_non_extended_date

    def _get_existing_indicators_sheets(self) -> Set['openpyxl.worksheet.worksheet.Worksheet']:
        """ Return a set of all existing worksheets from self._workbook, whose
        names correspond to any of the names from self._worksheet_properties.

        :return: Set of worksheet objects.
        """

        name_to_code = {
            self.__class__._worksheet_properties[code]['name']: code
            for code in self.__class__._worksheet_properties
        }
        indicators_worksheets = set()

        for sheet_name in self._workbook.sheetnames:
            try:
                code = name_to_code[sheet_name]
            except KeyError:
                continue
            else:
                indicators_worksheets.add(self._create_sheet(code))

        return indicators_worksheets

    def _protect_all_sheets(self) -> None:
        """ Protect all sheets described in self.__class__._worksheet_properties,
        using the parameters and values from self.__class__._worksheet_protection.

        :return: None.
        """

        for worksheet in self._get_existing_indicators_sheets():
            for property_, value in self.__class__._worksheet_protection.items():
                setattr(worksheet.protection, property_, value)

    @utils.log_func_time(logger, 20)
    def save(self) -> None:
        """ Save self._workbook at self._workbook_path."""

        logger.info(f'Saving workbook on: {self._workbook_path}')

        self._metadata_writer.write_indicators_last_date()

        logging.info('Protecting sheets')
        self._protect_all_sheets()

        self._workbook.save(self._workbook_path)
